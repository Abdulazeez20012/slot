"""Export functionality for member data."""

import csv
import json
from abc import ABC, abstractmethod
from datetime import datetime
from pathlib import Path
from typing import Any

from .models import ScrapeResult, TelegramMember


class BaseExporter(ABC):
    """Base class for member data exporters."""
    
    @abstractmethod
    def export(
        self,
        members: list[TelegramMember],
        output_path: Path,
        **kwargs,
    ) -> Path:
        """Export members to a file."""
        pass
    
    def _member_to_dict(self, member: TelegramMember) -> dict[str, Any]:
        """Convert a member to a dictionary for export."""
        return {
            "user_id": member.user_id,
            "username": member.username or "",
            "first_name": member.first_name,
            "last_name": member.last_name or "",
            "display_name": member.display_name,
            "phone": member.phone or "",
            "status": member.status.value,
            "last_seen": member.last_seen.isoformat() if member.last_seen else "",
            "is_bot": member.is_bot,
            "is_premium": member.is_premium,
            "is_verified": member.is_verified,
        }


class CSVExporter(BaseExporter):
    """Export members to CSV format."""
    
    def export(
        self,
        members: list[TelegramMember],
        output_path: Path,
        **kwargs,
    ) -> Path:
        """Export members to CSV file."""
        output_path = output_path.with_suffix(".csv")
        
        if not members:
            output_path.write_text("")
            return output_path
        
        fieldnames = [
            "user_id", "username", "first_name", "last_name", 
            "display_name", "phone", "status", "last_seen",
            "is_bot", "is_premium", "is_verified"
        ]
        
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for member in members:
                writer.writerow(self._member_to_dict(member))
        
        return output_path


class JSONExporter(BaseExporter):
    """Export members to JSON format."""
    
    def export(
        self,
        members: list[TelegramMember],
        output_path: Path,
        pretty: bool = True,
        **kwargs,
    ) -> Path:
        """Export members to JSON file."""
        output_path = output_path.with_suffix(".json")
        
        data = {
            "exported_at": datetime.now().isoformat(),
            "total_members": len(members),
            "members": [self._member_to_dict(m) for m in members],
        }
        
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2 if pretty else None, ensure_ascii=False)
        
        return output_path


class TXTExporter(BaseExporter):
    """Export member IDs to plain text (one per line)."""
    
    def export(
        self,
        members: list[TelegramMember],
        output_path: Path,
        include_username: bool = False,
        **kwargs,
    ) -> Path:
        """Export member IDs to text file."""
        output_path = output_path.with_suffix(".txt")
        
        lines = []
        for member in members:
            if include_username and member.username:
                lines.append(f"{member.user_id}\t@{member.username}")
            else:
                lines.append(str(member.user_id))
        
        output_path.write_text("\n".join(lines), encoding="utf-8")
        return output_path


class XLSXExporter(BaseExporter):
    """Export members to Excel XLSX format."""
    
    def export(
        self,
        members: list[TelegramMember],
        output_path: Path,
        sheet_name: str = "Members",
        **kwargs,
    ) -> Path:
        """Export members to Excel file."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            )
        
        output_path = output_path.with_suffix(".xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Headers with styling
        headers = [
            "User ID", "Username", "First Name", "Last Name",
            "Display Name", "Phone", "Status", "Last Seen",
            "Is Bot", "Is Premium", "Is Verified"
        ]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        for row, member in enumerate(members, 2):
            data = self._member_to_dict(member)
            ws.cell(row=row, column=1, value=data["user_id"])
            ws.cell(row=row, column=2, value=data["username"])
            ws.cell(row=row, column=3, value=data["first_name"])
            ws.cell(row=row, column=4, value=data["last_name"])
            ws.cell(row=row, column=5, value=data["display_name"])
            ws.cell(row=row, column=6, value=data["phone"])
            ws.cell(row=row, column=7, value=data["status"])
            ws.cell(row=row, column=8, value=data["last_seen"])
            ws.cell(row=row, column=9, value=data["is_bot"])
            ws.cell(row=row, column=10, value=data["is_premium"])
            ws.cell(row=row, column=11, value=data["is_verified"])
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        wb.save(output_path)
        return output_path


def get_exporter(format: str) -> BaseExporter:
    """Get the appropriate exporter for a format."""
    exporters = {
        "csv": CSVExporter(),
        "json": JSONExporter(),
        "txt": TXTExporter(),
        "xlsx": XLSXExporter(),
        "excel": XLSXExporter(),
    }
    
    format = format.lower()
    if format not in exporters:
        raise ValueError(
            f"Unknown export format: {format}. "
            f"Supported: {', '.join(exporters.keys())}"
        )
    
    return exporters[format]


def export_result(
    result: ScrapeResult,
    output_path: Path,
    format: str = "csv",
    **kwargs,
) -> Path:
    """Export a scrape result to a file."""
    exporter = get_exporter(format)
    return exporter.export(result.members, output_path, **kwargs)
